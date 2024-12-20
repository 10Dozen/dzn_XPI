import openpyxl
from .item_config import ItemConfig


class XLSXReader:
    '''Wrapper for openpyxl providing methods'''
    MAPPING_FRAME_NAME = "Maps"

    def __init__(self, filename: str):
        self.workbook = openpyxl.load_workbook(filename)
        self.mapping = {}
        self.items = {}

        self.__read_mapping()
        for sheet in self.workbook.sheetnames:
            if sheet == self.MAPPING_FRAME_NAME:
                continue

            self.__read_items(sheet)

    def get_items(self) -> dict[str, list[ItemConfig]]:
        '''Returns parsed items'''
        return self.items

    def __read_mapping(self):
        frame = self.workbook[self.MAPPING_FRAME_NAME]
        current_mapping = None
        for par_idx in range(frame.max_column // 2):
            for row_idx, row in enumerate(
                frame.iter_rows(
                    min_col=2*par_idx+1,
                    max_col=2*par_idx+2
                )
            ):
                # -- Get name of the parameter from first row
                if row_idx == 0:
                    current_mapping = {}
                    self.mapping[row[0].value] = current_mapping
                    continue

                if not row[0].value:
                    continue

                # -- Map param to it's definition
                current_mapping[row[0].value] = row[1].value

    def __read_items(self, page: str):
        items = []

        frame = self.workbook[page]
        self.items[page.lower()] = items

        # -- Read cols names
        cols = None
        for row in frame.iter_rows(max_row=1):
            cols = [row[i].value for i in range(frame.max_column)]

        # -- Read rows data
        for row in frame.iter_rows(min_row=2):
            item = {}
            # -- Skip rows where XPI.Bundle is not defined
            if not row[0].value:
                continue

            item = self.__parse_item(
                {col: row[i].value for i, col in enumerate(cols)}
            )
            items.append(item)

    def __parse_item(self, raw_item_data: dict[str, str]):
        '''Converts raw data into an ItemConfig class,
        applying mapping rules
        '''
        # -- Apply mapping
        for k, v in raw_item_data.items():
            if not v or (k not in self.mapping):
                continue
            raw_item_data[k] = self.mapping[k][v]


        # -- Compose data
        classname = raw_item_data["Class"]
        if raw_item_data["Suffix"]:
            classname = f"{classname}_{raw_item_data['Suffix']}"

        pointer: tuple[str] | None = None
        if raw_item_data["Pointer Setting"]:
            pointer = (
                raw_item_data["Pointer Setting"],
                self.mapping["Pointer.Pos"][raw_item_data["Pointer.PosDir"]],
                self.mapping["Pointer.Dir"][raw_item_data["Pointer.PosDir"]],
            )

        flashlight: tuple[str] | None = None
        if raw_item_data["Flashlight Setting"]:
            flashlight = (
                raw_item_data["Flashlight Setting"],
                self.mapping["Flashlight.Pos"][raw_item_data["Flashlight.PosDir"]],
                self.mapping["Flashlight.Dir"][raw_item_data["Flashlight.PosDir"]],
            )

        return ItemConfig(
            classname=classname,
            parent_classname=raw_item_data["Parent"],
            bundle=raw_item_data["XPI.Bundle"],
            scope=raw_item_data["Scope"],
            item_type=raw_item_data['XPI.Type'],
            mode=raw_item_data['XPI.Mode'],
            asgd_type=int(raw_item_data['ASDG Type']) if raw_item_data['ASDG Type'] else None,
            pointer_data=pointer,
            flashlight_data=flashlight
        )
